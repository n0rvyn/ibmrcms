from os import system
from os import path
from datetime import date
import re
import subprocess
import platform
import time
from openpyxl import load_workbook
from openpyxl import Workbook

# Specifies the time that s3270 will wait for the host connection to complete.
timeout = '10'
# Specifies a different TCP port to connect to.
# port_num = '2021'
# Turns on data steam and event tracing by default
trace_on = True
# The name & position of trace file
# Log file will be written in directory the same as 'ibmPCommTLS.py'
# currDir = path.abspath(path.join(path.dirname(__file__), 'log'))
trace_dir = path.abspath(path.dirname('__file__'))
# Specifies a name for tracing
trace_file = 'tn3270.log'
trace_file_size = '10M'
# Sometime after command executed successfully,
# keyboard is unlocked, but input field remain protected,
# so increase the value of 'unlockDelay'
# to wait input field unprotected.
unlockDelayTime = '300'
os_type = platform.system()
# Turn on or off for auto reconnect
# Only effective for Linux&MacOS
# For Windows, '-xrm ws3270.reconnect: True' is just fine.
# Function not supported for now.
autoReconnect = False
limitTimes = 5
# Define UT Code & Non UT Code
coreUTCode = ['1010', '1048', '11', '12', '14', '15',
              '1711', '1712', '1713',
              '20', '24',
              '3030', '3040', '3048',
              '3131', '3148', '32',
              '4010', '4011', '4012', '41',
              '4473', '48', '73', '74', '75']
nonUTCode = ['05', '34', '4020', '4211', '4212', '47', '60', '80',
             '86', '90', '91', '93', '95']


def debugPrint(str2prt, color='white'):
    if color == 'red':
        strDebug = '\033[0;31mError: \033[0m'
    elif color == 'yellow':
        strDebug = '\033[0;33mWarning: \033[0m'
    elif color == 'cyan':
        strDebug = '\033[0;36mInfo: \033[0m'
    elif color == 'white':
        strDebug = '\033[0;37mMsg: \033[0m'
    elif color == 'blue':
        strDebug = '\033[0;34mDebug: \033[0m'
    elif color == 'green':
        strDebug = '\033[0;32mDebug: \033[0m'
    else:
        strDebug = '\033[0;35mDebug: \033[0m'
    print('{}{}'.format(strDebug, str2prt))


# trans list like [[a, b, c], [d, e, f]] to excel file.
def dict2excel(multiDataList, fileDir='.', fileName='listSaved.xlsx'):
    wb = Workbook()
    # ws = wb.create_sheet('RCMS Reporting')
    ws = wb.active

    startRow = 1
    for subList in multiDataList:
        startCol = 1
        for key in subList:
            ws.cell(row=startRow, column=startCol, value=key)
            startCol += 1
        startRow += 1

    wb.save(fileDir + '/' + fileName)
    return True


class emulator(object):
    def __init__(self, args, host, backupHost):
        # Specifies a timeout, in seconds, that x3270 will wait for a host connection to complete.
        # If this is not set, the timeout will be determined by the operating system.
        self.host = host
        self.backupHost = backupHost
        self.args = args
        self.cmd_str = ''
        self.data = []
        self.fieldProtectIgnore = False

        # If trace_on = False
        # Data steam and event tracing will be turned off
        if not trace_on:
            self.args[10:13] = ''

        try:
            self.socket = subprocess.Popen(self.args,
                                           stdout=subprocess.PIPE,
                                           stdin=subprocess.PIPE,
                                           stderr=None)
        # except RuntimeError:
        except FileNotFoundError:
            debugPrint('S3270 not found! Install via Homebrew, '
                       'Gentoo-prefix or source code.\n'
                       'http://x3270.bgp.nu', color='red')
            exit(128)
        else:
            debugPrint('Subprocess open successfully.')

        self.emu_stat = {'keyboard': 'L', 'screen': 'U', 'field': 'U', 'conn': 'N',
                               # 'num_of_rows': '', 'num_of_cols': '',
                               # 'cursor_row': '', 'cursor_col': '',
                               # 'windows_id': '', 'cmd_exec_time': '',
                               'result': 'ok'}
        '''
        ### 1. Keyboard State
        # If the keyboard is unlocked, the letter U.
        # If the keyboard is locked, or not connected, the letter L.
        # If the keyboard is locked because of an operator error
        # like field overflow, protected field, etc., the letter E.
        ### 2. Screen Formatting
        # If the screen is formatted, the letter F, unformatted or in NVT mode, the letter U.
        ### 3. Field Protection
        # If the field containing the cursor is protected, the letter P. Unprotected or unformatted, the letter U.
        ### 4. Connection State
        # If connected to a host, the string C(hostname). Otherwise, the letter N.
        ### 9. Cursor Row; The current cursor row (zero-origin).
        ### 10. Cursor Column; The current cursor column (zero-origin).
        ### 13. Command Execution State
        # After command been executed, the first line of output is status of emulator,
        # the second line is execution state of previous command.
        # For now, 'ok' stands for command been executed fine,
        # 'error' stands for command not found or wrong format.
        '''

    def str2stdin(self, str2send):
        try:
            str2send = str2send.encode() + b'\n'
        except UnicodeEncodeError:
            debugPrint('"utf-8" codec cannot encode characters: '
                       'surrogates not allowed!', color='red')
        self.socket.stdin.write(str2send)
        self.socket.stdin.flush()

    def read1line(self):
        return self.socket.stdout.readline().decode().rstrip('\n')

    def resetCursor(self):
        # When 'MoveCursor' to a field can not be edit,
        # keyboard return locked, field be protected.
        # Then send 'reset'|'home' to unlock keyboard.
        # self.socket.stdin.write('reset'.encode() + b'\n')
        # self.socket.stdin.flush()
        self.str2stdin('home')
        # 1st read line from stdout for Emulator Status.
        self.read1line()
        # 2nd read line from stdout for Command Executed Status.
        self.read1line()

    # send command to emulator socket, return emulator status.
    def exeCmd(self, str2send):
        self.str2stdin(str2send)
        while True:
            strLine = self.read1line()
            if not strLine.startswith('data'):
                self.emu_stat['keyboard'] = strLine.split()[0]
                self.emu_stat['screen'] = strLine.split()[1]
                self.emu_stat['field'] = strLine.split()[2]
                self.emu_stat['conn'] = strLine.split()[3]
                self.emu_stat['result'] = self.read1line()
                break
        return self.emu_stat

    def exeNewline(self):
        time.sleep(0.2)
        return self.exeCmd('')

    # if host connected, return True.
    def isConnected(self):
        if self.emu_stat['conn'].startswith('C'):
            return True
        else:
            return False
        
    # if emulator screen been formatted, return True.
    def isFormatted(self):
        if self.emu_stat['screen'] == 'F':
            return True
        else:
            return False

    # if emulator input field unprotected, return True.
    def isUnprotected(self):
        if self.emu_stat['field'] == 'U':
            return True
        else:
            return False

    def connect(self, backup=False):
        hostAddr = self.host.split(':')[0]
        backupHostAddr = self.backupHost.split(':')[0]
        if os_type == 'Windows':
            if system('ping -n 1 {}'.format(hostAddr)) != 0:
                debugPrint('Host not reachable!', color='red')
                return False
        elif system('ping -c 1 {} >/dev/null 2>&1'.format(hostAddr)) != 0:
            debugPrint('Host not reachable!', color='red')
            return False

        if self.isConnected():
            debugPrint('Host already connected! Disconnect firstly!')
            debugPrint('Disconnect Status: {}'.format(self.disconnect()))

        if not backup:
            print(self.exeCmd('connect(L:' + self.host + ')'))
        else:
            print(self.exeCmd('connect(L:' + self.backupHost + ')'))
        i = limitTimes
        while True:
            if self.isConnected():
                debugPrint('host {} connected.'.format(self.host))
                return True
            else:
                # print(self.exeNewline())
                self.exeNewline()
            i -= 1
            if i < 0:
                debugPrint('Connect timeout.')
                return False

    def disconnect(self):
        return self.exeCmd('disconnect')

    def quit(self):
        self.socket.stdin.write('quit'.encode() + b'\n')
        self.socket.stdin.flush()
        return True

    def getCurrScreen(self):
        self.data = []
        i = limitTimes
        while True:
            if self.isFormatted():
                break
            else:
                # print(self.exeNewline())
                self.exeNewline()
            i -= 1
            if i < 0:
                debugPrint('Screen can not be formatted! Host may not connected.')
                return False
        self.str2stdin('printtext(string)')
        while True:
            readl = self.read1line()
            # debugPrint(line)
            if readl.startswith('data'):
                self.data.append(readl.replace('data: ', '', 1))
                if readl == 'data:  No LUs available now, please try later':
                    debugPrint('Error detected! Reset data field. Try to reconnect.')
                    self.data = []
                    self.connect(backup=True)
                    debugPrint('help!!!')
            else:
                self.read1line()
                break
        if self.data:
            return True
        else:
            return False

    def printCurrScreen(self):
        if self.getCurrScreen():
            for dataLine in self.data:
                print(dataLine)
            return True
        else:
            return False

    def getString(self, row, col, length):
        self.getCurrScreen()
        if not self.data:
            return ''
        else:
            return self.data[row-1][col-1: col-1+length]

    def sendStringDirectly(self, str2send):
        i = limitTimes
        while i > 0:
            # if self.isUnprotected():
            if self.isUnprotected() and self.isFormatted():
                return self.exeCmd('string("{}")'.format(str2send))
            else:
                self.exeCmd('home')
                self.exeNewline()
            i -= 1
        return False

    def sendString(self, row, col, str2send):
        self.exeCmd('MoveCursor({},{})'.format(row-1, col-1))
        return self.sendStringDirectly(str2send)

    # Send Enter AID
    def sendEnter(self):
        return self.exeCmd('enter')

    # Program Function AID (n from 1 to 24)
    def sendPf(self, n):
        if 1 <= n <= 24:
            return self.exeCmd('pf({})'.format(n))
        else:
            return False

    # Send Program Attention AID (n from 1 to 3)
    def sendPA(self, n):
        if 1 <= n <= 3:
            return self.exeCmd('PA({})'.format(n))
        else:
            return False

    # Delete the entire field.
    def deleteField(self):
        return self.exeCmd('DeleteField')

    # Erase all input fields.
    def eraseInput(self):
        return self.exeCmd('EraseInput')


class x3270(emulator):
    def __init__(self, host, backupHost):
        # For MacOS & Linux, emulator is called 's3270'
        if os_type == 'Darwin' or os_type == 'Linux':
            self.args = ['s3270', '-model', '3279-2',
                         '-xrm', 's3270.charset: chinese-gb18030',
                         '-xrm', 's3270.unlockDelay: True',
                         # Increase delay time to wait field protection be letter 'U' not 'P'
                         # So strings can be send by next 'string(str2send)' command.
                         '-xrm', 's3270.unlockDelayMs: ' + unlockDelayTime,
                         # x3270 will automatically reconnect to a host after it disconnects
                         # not support by x3270 macOS version
                         # '-xrm', 's3270.reconnect: True',
                         # port specified is moved to format like: 1.1.1.1:992
                         # '-port', port_num,
                         '-utf8',
                         # '-trace', '-tracefile', '>>' + trace_dir + '/' + trace_file,
                         '-trace', '-tracefile', trace_dir + '/' + trace_file,
                         '-tracefilesize', trace_file_size,
                         '-noverifycert',
                         '-connecttimeout', timeout]
        # For Windows, emulator is 'wc3270'
        # Will be supported later.
        elif os_type == 'Windows':
            self.args = ['ws3270', '-model', '3279-2',
                         '-xrm', 'ws3270.charset: chinese-gb18030',
                         '-xrm', 'ws3270.unlockDelay: True',
                         '-xrm', 'ws3270.unlockDelayMs: ' + unlockDelayTime,
                         # x3270 will automatically reconnect to a host after it disconnects
                         # '-xrm', 'ws3270.reconnect: True',
                         # port specified is moved to format like: 1.1.1.1:992
                         # '-port', port_num,
                         '-utf8',
                         # '-trace', '-tracefile', '>>' + trace_dir + '\\' + trace_file,
                         '-trace', '-tracefile',  trace_dir + '\\' + trace_file,
                         '-tracefilesize', trace_file_size,
                         '-noverifycert',
                         '-connecttimeout', timeout]
        else:
            raise OSError('Unknown OS!')

        emulator.__init__(self, self.args, host, backupHost)


class rcms(x3270):
    def __init__(self, host, backupHost, ssrSN, passwd):
        self.host = host
        self.backupHost = backupHost
        # host is '[host address]:[port]' format
        # self.host = host.split(':')[0]
        # self.backupHost = backupHost.split(':')[0]
        x3270.__init__(self, self.host, self.backupHost)
        self.userid = 'C' + ssrSN.upper()
        self.passwd = passwd
        self.ssrSN = ssrSN
        self.callQueue = ['SDLNCN', 'SDLINC']
        self.callType = ['CRR', 'INS', 'PMA']
        self.callStatus = ['cls', 'pak']
        self.callCE = ''
        self.rcmsData = []
        self.currCallNo = ''
        # RCMS special pages & keyword
        self.pages = {'ACDN': 'First page after connected.',
                      # 24, 1, 4
                      'LOGON': 'RCMS login interface.',
                      # 23, 10, 5
                      'CMD=>': 'Pages has "CMD" prompt.',
                      # 22, 2, 5
                      'employee': 'Report entry.',
                      # 6, 25, 8
                      'report': 'CE reporting list.',
                      # 4, 64, 6
                      'ACTION': 'RCMS action plan.',
                      # 1, 23, 6
                      'PF4': 'Press PF4 to exit cab.',
                      # 7, 10, 3
                      'READ': 'Action plan - read only mode. Enter to continue.'
                      # 1, 37, 4
                      }

    def currPageType(self):
        keyword = ''
        if self.getString(24, 1, 4) == 'ACDN':
            keyword = 'ACDN'
        elif self.getString(23, 10, 5) == 'LOGON':
            keyword = 'LOGON'
        elif self.getString(1, 37 ,4) == 'READ':
            keyword = 'READ'
        elif self.getString(22, 2, 5) == 'CMD=>':
            keyword = 'CMD=>'
        elif self.getString(6, 25, 8) == 'employee':
            keyword = 'employee'
        elif self.getString(4, 64, 6) == 'report':
            keyword = 'report'
        elif self.getString(1, 23, 6) == 'ACTION':
            keyword = 'ACTION'
        elif self.getString(7, 10, 3) == 'PF4':
            keyword = 'PF4'

        if keyword in self.pages.keys():
            return self.pages[keyword]

    # return True if login successfully, else return False.
    def login(self):
        """
        # Moved to method go2MainMenu()
        if not self.isConnected():
            self.connect()
        if not self.isConnected():
            self.connect(backup=True)
        if self.currPageType() == self.pages['CMD=>']:
            debugPrint('Host already connected & login.')
            return True
        self.sendStringDirectly('rcms')
        self.sendEnter()
        # self.printCurrScreen()
        if self.currPageType() != self.pages['LOGON']:
            debugPrint('Wrong page, login denied.')
            return False
        """
        # Erase all input field.
        self.eraseInput()
        if len(self.userid) != 7:
            debugPrint('Wrong userid!')
            return False
        self.sendString(18, 30, self.userid)
        if len(self.passwd) > 23:
            debugPrint('Password too long!')
            return False
        # print('passwd: {}'.format(self.passwd))
        self.sendString(18, 54, self.passwd)
        self.sendEnter()
        prompt = self.getString(24, 11, 60).strip()
        if prompt == 'USERID NOT DEFINED IN RACF':
            debugPrint('Wrong userid, check please.', color='red')
            exit(129)
        if prompt == 'YOUR PASSWORD WAS NOT SUPPLIED AND IS REQUIRED.':
            debugPrint('No password! Return to welcome page, '
                       'in case of userid been revoked!', color='red')
            self.sendPf(3)
            return False
        if prompt == 'THE PASSWORD YOU SUPPLIED IS WRONG.':
            debugPrint('Wrong password, In case of userid been revoked, '
                       'return to welcome page.', color='red')
            self.sendPf(3)
            return False
        if prompt == 'YOUR PASSWORD HAS EXPIRED. PLEASE TYPE YOUR NEW PASSWORD':
            debugPrint('Password expired, change via HOD or IBM PCOmm.', color='red')
            self.sendPf(3)
            exit(129)
        # time.sleep(5)
        """
        if self.currPageType() == self.pages['CMD=>']:
            return True
        else:
            return False
        """

    def go2MainMenu(self):
        """
        double sendPf(6) --> double sendPf(4)
        if current page is first page; then type "rcms"
        if current page is login page; login with userid & passwd
        if current page is other page; back2 main menu
        check if there is 'CMD' prompt in current page
        :return: there is 'CMD' return True or return False
        """
        if not self.isConnected():
            debugPrint('Host not connected! Attempt to reconnect.', color='red')
            self.connect()
        if not self.isConnected():
            debugPrint('Reconnect failed, try backup host.', color='red')
            self.connect(backup=True)
        if not self.isConnected():
            debugPrint('Both hosts connect failed. Abort!', color='red')
            return False
        # debugPrint('Connected. Check current page type.', color='green')
        if self.getString(22, 1, 30) == 'ENTER Application Name or MENU':
            # debugPrint('First page after connected!', color='green')
            self.sendString(23, 1, 'RCMS')
            self.sendEnter()
        if self.getString(14, 22, 23) == 'CUSTOMER SERVICE SYSTEM':
            # debugPrint('Login interface.', color='green')
            self.login()
        if self.getString(22, 2, 5) != 'CMD=>':
            # debugPrint('No "CMD" on current page, try back to main menu.', color='yellow')
            self.sendPf(6)
            self.sendPf(6)
            self.sendPf(4)
            self.sendPf(4)
        if self.getString(22, 2, 5) == 'CMD=>':
            self.eraseInput()
            self.sendString(22, 8, 'MENU')
            self.sendEnter()
            # debugPrint('Page has "CMD" prompt. command is permitted.', color='green')
            return True
        debugPrint('Unexpected Error! Return False.', color='red')
        return False

    def nextPage(self, n):
        return self.sendPf(n)

    def getMultiPages(self, key2stop, key2stopRow, key2stopCol, lenOfKey2stop,
                      row2start, col2start, row2stop, col2stop, PFN=8):
        self.rcmsData = []
        i = 0
        while True:
            dataLine = self.getString(row2start+i, col2start, col2stop-col2start)
            if dataLine:
                self.rcmsData.append(dataLine)
            if i > row2stop - row2start - 1:
                self.nextPage(PFN)
                i = 0
            else:
                i += 1
            if self.getString(key2stopRow, key2stopCol, lenOfKey2stop).strip() == key2stop:
                break
        return self.rcmsData

    def callRead(self, callNo=''):
        callInfo = {'customer': '', 'contact': '',
                    'contract end date': '', 'RCMS call': '', 'last call': '',
                    'problem type': '', 'prod/model sn': '',
                    # 'prod': '\t', # 'model': '\t',# 'sn': '\t',
                    'desc': '\t', 'type': '\t', 'st': '\t', 'sev': '\t',
                    'recall': '', 'ass. CE': '', 'create time': '',
                    'update time': '', 'sc/ac': '', 'msg': '',
                    # 'sc': '', # 'ac': '',
                    'type of maintenance': '', 'service level': '',
                    'comments': '\n', 'ecurep': '\n', 'action plan': '\n'}
        if callNo:
            self.currCallNo = callNo
        elif not self.currCallNo:
            debugPrint('No last call, nothing got.')
            return callInfo

        """
        if self.currPageType() != self.pages['CMD=>']:
            return callInfo
        """
        self.go2MainMenu()
        if self.getString(23, 11, 22) == 'Call record not found.':
            debugPrint('Wrong call no., nothing got.')
            return callInfo
        debugPrint('Read {} in processing, wait for seconds...'
                   .format(self.currCallNo.upper()), color='green')
        self.sendString(22, 8, 'cr,' + self.currCallNo)
        # self.sendString(22, 8, 'cr,' + callNo)
        self.sendEnter()
        callInfo['contract end date'] += self.getString(9, 34, 8)
        custNo = self.getString(6, 24, 6)
        self.sendEnter()
        callInfo['customer'] += '{} ({})'.format(self.getString(4, 15, 18).strip(), custNo)
        # callInfo['customer num'] += self.getString(4, 57, 6)
        telephone = self.getString(5, 16, 20)
        callInfo['RCMS call'] += self.getString(3, 59, 7)
        callInfo['last call'] += self.getString(5, 59, 18)
        callInfo['contact'] += '{} {}'.format(self.getString(8, 15, 25).strip(), telephone)
        callInfo['problem type'] += self.getString(9, 16, 1)
        callInfo['prod/model sn'] += self.getString(10, 8, 4) + '/' \
                                     + self.getString(10, 24, 3) + ' ' \
                                     + self.getString(10, 32, 7)
        # callInfo['prod'] += self.getString(10, 8, 4)
        # callInfo['model'] += self.getString(10, 24, 3)
        # callInfo['sn'] += self.getString(10, 32, 7)
        callInfo['desc'] += self.getString(11, 8, 30).strip()
        callInfo['type of maintenance'] += self.getString(13, 2, 20).strip()
        callInfo['service level'] += self.getString(14, 2, 78).strip()
        callInfo['type'] += self.getString(15, 11, 3)
        callInfo['st'] += self.getString(15, 19, 3) + ' ' + self.getString(16, 61, 15).replace(' ', ':', 1)
        callInfo['sev'] += self.getString(15, 28, 1)
        callInfo['recall'] += self.getString(15, 37, 3).strip()
        callInfo['ass. CE'] += self.getString(16, 11, 30).strip()
        callInfo['create time'] += self.getString(17, 44, 15)
        callInfo['update time'] += self.getString(17, 61, 15)
        callInfo['msg'] += self.getString(12, 8, 30).strip()\
            .replace('JV MACHINE', '\033[0;31mJV MACHINE\033[0m')

        for i in range(0, 3):
            callInfo['comments'] += '{0}\n'.format(self.getString(19+i, 1, 80).rstrip().lstrip(' comments '))

        callContent = self.getMultiPages('Last page.', 23, 11, 10, 5, 2, 20, 79)
        fromIndex = 0
        toIndex = 0
        tempList = []

        while True:
            for cc in callContent:
                if '**ECUREP START**' in cc:
                    fromIndex = callContent.index(cc)
                if '**ECUREP END**' in cc:
                    toIndex = callContent.index(cc)
            if fromIndex == toIndex:
                break
            tempList.extend(callContent[fromIndex+1: toIndex])
            callContent[fromIndex] = ''
            callContent[toIndex] = ''
            fromIndex = 0
            toIndex = 0

        for tl in tempList:
            callInfo['ecurep'] += '{}\n'.format(tl)

        actionPlan = []
        self.sendPf(6)
        self.sendString(22, 8, 'ca,action')
        self.sendEnter()
        self.sendEnter()
        if self.getString(1, 47, 2) == '01':
            for i in range(8, 19):
                acLine = self.getString(i, 2, 79)
                if acLine:
                    actionPlan.append(acLine)
        elif self.getString(24, 6, 9) == 'Prev Plan':
            actionPlan = self.getMultiPages('FIRST', 23, 2, 5, 8, 2, 19, 79, 1)
        self.sendPf(6)
        self.sendPf(6)
        self.sendPf(6)
        self.sendString(22, 8, 'MENU')
        self.sendEnter()

        for ap in actionPlan:
            if ap.startswith('SC '):
                callInfo['sc/ac'] = ap[3:5] + ap[9:11]
            elif not ap.startswith('Error Code'):
                callInfo['action plan'] += '{}\n'.format(ap)

        returnStr = ''
        for ci in callInfo:
            returnStr += '\033[0;31m{}: \033[0m{}\n'.format(ci.upper(), callInfo[ci])
        print(returnStr)
        return callInfo

    def callSearch(self, ssrSN='', queue='', callType='', cust='', ce=''):
        # 23, 11, 'More pages available.'
        # 23, 11, 'Last page.'
        self.go2MainMenu()
        """
        if ssrSN:
            ce = self.transSN2notesID(ssrSN)
        """
        print(cust)
        if not ce and not cust and not ssrSN:
            ce = self.transSN2notesID(self.ssrSN)
        if not queue:
            queue = 'SDLNCN'
        self.sendString(22, 8, 'cs,'
                        + queue + ',,,,,'
                        + callType + ',,,,,,'
                        + cust + ',,'
                        + ce + ',A,0,,,')
        self.sendEnter()
        return self.getMultiPages('Last page.', 23, 11, 10, 5, 5, 17, 79)

    def callSrchPlus(self, argsList):
        srchArgs = {'notesID': '',
                    'custNo': '',
                    'callType': '',
                    'callQueue': '',
                    'ssrSN': ''}
        for args in argsList:
            typeOfArgs = self.typeOfArgs(args)[0]
            realValue = self.typeOfArgs(args)[1]
            for key in srchArgs:
                if typeOfArgs == key and not srchArgs[key]:
                    srchArgs[key] = realValue

        print(srchArgs)
        return self.callSearch(ssrSN=srchArgs['ssrSN'],
                               queue=srchArgs['callQueue'],
                               callType=srchArgs['callType'],
                               cust=self.transCustNo2Name(srchArgs['custNo']),
                               ce=srchArgs['notesID'])

    def callInquiry(self, custNo='', machSN='', machType='', fromDate='', toDate=''):
        if not custNo and not machSN and not machType:
            return []
        self.go2MainMenu()
        if len(custNo) not in [0, 6, 8]:
            debugPrint('Wrong customer No.!')
            return []
        else:
            self.sendString(22, 8, 'ci,' + custNo + ',' + machSN
                            + ',,,,,,,,' + fromDate + ',' + toDate)
            self.sendEnter()
            return self.getMultiPages('Last page.', 23, 11, 10, 5, 5, 18, 79)

    def callInquiryPlus(self, argsList):
        ciArgs = {'custNo': '',
                  'machSN': '',
                  'machType': ''}
        for args in argsList:
            typeOfArgs = self.typeOfArgs(args)[0]
            realValue = self.typeOfArgs(args)[1]
            for key in ciArgs:
                if typeOfArgs == key and not ciArgs[key]:
                    ciArgs[key] = realValue
        print(ciArgs)
        return self.callInquiry(custNo=ciArgs['custNo'],
                                machSN=ciArgs['machSN'],
                                machType=ciArgs['machType'])

    def transSN2notesID(self, ssrSN):
        """
        if self.currPageType() != self.pages['CMD=>']:
            return 'Unknown'
        """
        if not ssrSN:
            return ''
        self.go2MainMenu()
        self.sendString(22, 8, 'red,0' + ssrSN)
        self.sendEnter()
        if self.getString(23, 11, 14) == 'No C.E. found.':
            return ''
        return self.getString(4, 45, 18).rstrip()

    def transID2SN(self, ssrNotesID):
        if not ssrNotesID:
            return ''
        self.go2MainMenu()
        """
        if self.currPageType() != self.pages['CMD=>']:
            return 'Unknown'
        """
        self.sendString(22, 8, 'red,' + ssrNotesID)
        self.sendEnter()
        if self.getString(23, 11, 14) == 'No C.E. found.':
            return ''
        return self.getString(4, 38, 6)

    def transCustNo2Name(self, custNo):
        """
        if self.currPageType() != self.pages['CMD=>']:
            return 'Unknown'
        """
        custName = ''
        if not custNo:
            return custName
        self.go2MainMenu()
        self.sendString(22, 8, 'ci,' + custNo)
        self.sendEnter()
        if self.getString(5, 6, 1) == '1':
            self.sendStringDirectly('1')
            self.sendEnter()
            custName = self.getString(4, 2, 20)
        return custName.lstrip('account name: ').strip()

    def typeOfArgs(self, args):
        """
        20A         not support

        1814        machine type

        805543      cust no.
        avr71d      ssr sn

        0avr71d     ssr sn
        cavr71d     ssr sn
        0805543     cust no.
        78k15c8     machine sn

        00805543    cust no.
        zhi jie     notes id
        :return:
        """
        lenOf = len(args)
        args = args.upper()
        # debugPrint('Args input: \033[0;31m{}\033[0m, length of: {}'.format(args, lenOf))
        if args in self.callType:
            return ['callType', args]
        if args in self.callQueue:
            return ['callQueue', args]
        if lenOf == 4:
            # debugPrint('Maybe machine type.')
            if self.proInv(machType=args):
                return ['machType', args]
        if lenOf == 6:
            # debugPrint('Maybe cust no.')
            if self.transCustNo2Name(args):
                return ['custNo', args]
            # debugPrint('Maybe ssr sn.')
            if self.transSN2notesID(args):
                return ['ssrSN', args]
        if lenOf == 7:
            # debugPrint('Maybe machine sn.')
            if self.proInv(machSN=args):
                return ['machSN', args]
            # debugPrint('Maybe cust no. starts with "0".')
            if self.transCustNo2Name(args[1:7]):
                return ['custNo', args[1:7]]
            # debugPrint('Maybe ssr sn starts with "C" or "0".')
            if self.transSN2notesID(args[1:7]):
                return ['ssrSN', args[1:7]]
        if lenOf == 8:
            # debugPrint('Maybe cust no. starts with "00"')
            if self.transCustNo2Name(args[2:8]):
                return ['custNo', args[2:8]]
        # debugPrint('Verify if it is notes ID.')
        if self.transID2SN(args):
            return ['notesID', args]
        return ['Unknown', args]

    def transCusName2No(self, custName):
        """
        usefully only the customer specified has at least one call opened.
        :param custName:
        :return:
        """
        custNo = ''
        if not custName:
            return custNo
        self.go2MainMenu()
        self.sendString(22, 8, 'CS,SDLNCN,,,,,,,,,,,' + custName + ',,,A,0,,,')
        self.sendEnter()
        if self.getString(5, 6, 1) == '1':
            self.sendStringDirectly('1')
            self.sendEnter()
            custNo = self.getString(6, 24, 6).strip()
        if not custNo:
            debugPrint('No call in "SDLNCN" queue, try "SDLINC" for IPS call.')
            self.go2MainMenu()
            self.sendString(22, 8, 'CS,SDLINC,,,,,,,,,,,' + custName + ',,,A,0,,,')
            self.sendEnter()
            if self.getString(5, 6, 1) == '1':
                self.sendStringDirectly('1')
                self.sendEnter()
                custNo = self.getString(6, 24, 6).strip()
        if not custNo:
            debugPrint('The customer name you specified has no call open, '
                       'transfer name to no failed.')
        return custName + ': ' + custNo

    def proInv(self, machType='', machSN='', custNo=''):
        """
        if self.currPageType() != self.pages['CMD=>']:
            return ''
        """
        if not machType and not machSN and not custNo:
            return []
        self.go2MainMenu()
        self.sendString(22, 8, 'in,' + machType + ',' + machSN + ',,' + custNo)
        # 'IN,1814,78K15C8,,805543'
        self.sendEnter()
        return self.getMultiPages('Last page.', 23, 11, 10, 5, 6, 21, 78)

    def proInvPlus(self, argsList):
        inArgs = {'custNo': '',
                  'machSN': '',
                  'machType': ''}
        for args in argsList:
            typeOfArgs = self.typeOfArgs(args)[0]
            realValue = self.typeOfArgs(args)[1]
            for key in inArgs:
                if typeOfArgs == key and not inArgs[key]:
                    inArgs[key] = realValue
        print(inArgs)
        return self.proInv(machType=inArgs['machType'],
                           machSN=inArgs['machSN'],
                           custNo=inArgs['custNo'])

    def proInvDetail(self):
        self.sendStringDirectly('1')
        self.sendEnter()
        return self.getMultiPages('Last page.', 23, 11, 10, 3, 2, 20, 78)

    def readTimeReport(self, ssrSN, fromDate, toDate, repLstPrt=True,
                       save=False, save2Dir='', save2file='',
                       utSum=False, prtDetail=False):
        if not save2file:
            save2file = ssrSN + '_' +fromDate + '_' + toDate + '.xlsx'
        if not save2Dir:
            save2Dir = './log'

        detailSum = []
        detailTemp = []
        reportList = []
        reportList4UT = []
        scHours = {'1010': 0, '11': 0, '12': 0, '14': 0, '1048': 0,
                   '15': 0, '17': 0, '20': 0, '24': 0, '1712': 0, '1711': 0,
                   '30': 0, '31': 0, '32': 0, '4010': 0,
                   '4011': 0, '4012': 0, '41': 0, '4473': 0,
                   '48': 0, '73': 0, '74': 0, '75': 0,
                   '05': 0, '34': 0, '4020': 0, '4211': 0, '4212': 0,
                   '44': 0, '47': 0, '60': 0, '80': 0,
                   '86': 0, '90': 0, '91': 0, '93': 0, '95': 0}
        detailItem = {'reported day': '',
                      'RCMS call': '',
                      'employee': '',
                      'bu': '',
                      'sc': '',
                      'ac': '',
                      'cust no': '',
                      'cust name': '',
                      'mt or sn': '',
                      'total': '',
                      'travel': '',
                      'stop': '',
                      'mach st': '',
                      'hours day': '',
                      'week': '',
                      'hours week': '',
                      'report': '',
                      'status': '',
                      'force': '',
                      'log': '',
                      'holiday': '',
                      'cc': '',
                      'tvl exp': '',
                      'pause': '',
                      'pw': '',
                      'otd': '',
                      'otc': '',
                      'tom': '',
                      'bill sys msg': '',
                      'desc': '',
                      'note': ''}
        self.go2MainMenu()
        """
        if self.currPageType() != self.pages['CMD=>']:
            return []
        """
        for key in detailItem:
            detailTemp.append(key)
        detailSum.append(detailTemp)
        fromYear = fromDate[0:2]
        fromMonth = fromDate[2:4]
        fromDay = fromDate[4:6]
        toYear = toDate[0:2]
        toMonth = toDate[2:4]
        toDay = toDate[4:6]
        redDate = fromDay + fromMonth + fromYear
        # fromDateKey = fromYear + '/' + fromMonth + '/' + fromDay
        toDateKey = toYear + '/' + toMonth + '/' + toDay
        # print('fromDay: ', fromDay, 'toDay: ', toDay)
        # avaHrs = (0 - ((int(fromDay) - 30) - int(toDay))) // 7 * 40
        avaHrs = (date(int(toYear), int(toMonth), int(toDay))
                  - date(int(fromYear), int(fromMonth), int(fromDay))).days // 7 * 40
        self.sendString(22, 8, 'red,0' + ssrSN + ',' + redDate)
        self.sendEnter()
        # self.sendPf(12)
        while True:
            # After read reporting list, if the day is the end day, then quit loop.
            self.sendPf(1)
            # self.printCurrScreen()
            prompt=self.getString(4, 7, 18)
            print('Day: ' + prompt, end='\r', flush=True)
            reportList.append('\033[0;31m{}\033[0m'
                              .format(self.getString(4, 7, 20).strip()))
            i = 0
            while True:
                reportRow = self.getString(7 + i, 5, 2).strip()
                if not reportRow:
                    break
                reportList.append(self.getString(7 + i, 5, 79))
                reportList4UT.append(self.getString(7 + i, 5, 79))
                # self.sendString(22, 8, int(reportRow))
                self.sendString(22, 8, reportRow)
                self.sendEnter()
                detailItem['reported day'] = self.getString(7, 2, 15)
                detailItem['hours day'] = self.getString(3, 48, 3)
                detailItem['week'] = self.getString(3, 60, 2)
                detailItem['hours week'] = self.getString(3, 75, 3)
                detailItem['report'] = self.getString(4, 10, 9)
                detailItem['status'] = self.getString(4, 28, 3)
                detailItem['force'] = self.getString(4, 40, 1)
                detailItem['log'] = self.getString(4, 48, 1)
                detailItem['RCMS call'] = self.getString(4, 61, 7)
                detailItem['employee'] = self.getString(7, 20, 28).rstrip()
                detailItem['bu'] = self.getString(7, 49, 8)
                detailItem['sc'] = self.getString(10, 2, 2)
                detailItem['ac'] = self.getString(10, 6, 2)
                detailItem['cc'] = self.getString(10, 11, 1)
                detailItem['total'] = self.getString(10, 15, 3)
                detailItem['travel'] = self.getString(10, 24, 2)
                detailItem['stop'] = self.getString(10, 31, 3)
                detailItem['otd'] = self.getString(10, 63, 1)
                detailItem['otc'] = self.getString(10, 69, 1)
                detailItem['mt or sn'] = self.getString(13, 2, 15).replace(' ', '')
                detailItem['mach st'] = self.getString(13, 19, 1)
                detailItem['tom'] = self.getString(13, 26, 1)
                detailItem['cust no'] = self.getString(16, 4, 6)
                detailItem['cust name'] = self.getString(16, 12, 24)
                detailItem['bill sys msg'] = self.getString(18, 16, 15)
                detailItem['desc'] = self.getString(19, 7, 40).strip()
                detailItem['note'] = self.getString(20, 7, 40).strip()
                if detailItem['sc'] in ['80', '90', '95']:
                    detailItem['note'] = self.getString(13, 7, 40).strip()
                    detailItem['cust no'] = ''
                    detailItem['cust name'] = ''
                    detailItem['mt or sn'] = ''
                if detailItem['sc'] in ['75', '42']:
                    detailItem['cust no'] = self.getString(13, 4, 6)
                    detailItem['cust name'] = self.getString(13, 12, 24)
                    detailItem['mt or sn'] = ''
                    detailItem['bill sys msg'] = 'contract no: ' \
                                                 + self.getString(15, 63, 15) \
                                                 + '; work number: ' \
                                                 + self.getString(16, 68, 10).strip()
                detailTemp = []
                for di in detailItem:
                    detailTemp.append(detailItem[di])
                    # detailSum.append(detailItem[di])
                detailSum.append(detailTemp)
                self.sendPf(6)
                self.sendPf(4)
                i += 1
            print(' '*30, end='\r', flush=True)

            if self.getString(4, 7, 8) == toDateKey:
                break
        if save:
            dict2excel(detailSum, fileDir=save2Dir, fileName=save2file)
            print('File save to : "{}"'.format(save2Dir + '/' + save2file))
        if repLstPrt:
            for rl in reportList:
                print(rl)
        if prtDetail:
            detailSum.remove(detailSum[0])
            for ds in detailSum:
                print(ds)

        if utSum:
            coreUTHours = 0
            nonUTHours = 0
            coreUTSumList = {}
            nonUTSumList = {}

            for rl4ut in reportList4UT:
                if rl4ut:
                    sc = rl4ut[17:19]
                    ac = rl4ut[20:22].strip()
                    tot = rl4ut[8:11]
                    if not tot:
                        tot = 0
                    totHrs = int(tot) / 10
                    codeIndex = sc + ac
                    if codeIndex in coreUTCode:
                        coreUTHours += totHrs
                    elif codeIndex in nonUTCode:
                        nonUTHours += totHrs
                    else:
                        debugPrint('Code: "{}" is not defined.'.format(codeIndex))
                    try:
                        scHours[codeIndex] += totHrs
                    except KeyError:
                        debugPrint('Code Not Defined! Contact Author for Help.', color='yellow')

            for code in scHours:
                if scHours[code] != 0:
                    if code in coreUTCode:
                        coreUTSumList[code] = scHours[code]
                    elif code in nonUTCode:
                        nonUTSumList[code] = scHours[code]
            UTPercent = '{:>.2f}%'.format(coreUTHours * 100 / int(avaHrs))

            # print(' '*30, end='\r', flush=True)
            print('Core UT List: \t{}'.format(coreUTSumList))
            print('Non UT List: \t{}'.format(nonUTSumList))
            print('Core UT Hrs: \t{}'.format(coreUTHours))
            print('Non UT Hrs: \t{}'.format(nonUTHours))
            print('Avai Hours: \t{}'.format(avaHrs))
            print('UT percent: \t{}'.format(UTPercent))
        return detailSum

    def getUTSum4Team(self, ssrSNList, fromDate, toDate):
        for ssr in ssrSNList:
            print('\033[0;36mSSR "{}"\033[0m'.format(self.transSN2notesID(ssr)))
            self.readTimeReport(ssr, fromDate, toDate, repLstPrt=False, utSum=True)

    def teamEPSBCheck(self, ssrSNList, fromDate, toDate):
        epsbRepList = []
        machSnList = []
        machOverList = []
        epsbRepOverList = []
        debugPrint('Verify Reporting EPSB Overlap for SSR List:\n'
                   '{}'.format(ssrSNList))
        for ssr in ssrSNList:
            debugPrint('Collect Reporting Data For '
                       'SSR \033[0;36m{}\033[0m:'.format(ssr.upper()))
            for detailLine in self.readTimeReport(ssr, fromDate, toDate, repLstPrt=False):
                if '17' in detailLine and '12' in detailLine:
                    epsbRepList.append(detailLine[0:9])
            debugPrint('Work for SSR {} Done.'.format(ssr.upper()))
        for eplst in epsbRepList:
            machSnList.append(eplst[8])
        for msl in machSnList:
            if machSnList.count(msl) > 1:
                machOverList.append(msl)
        for eplst in epsbRepList:
            if eplst[8] in set(machOverList):
                eplst.append('Overlap')
                epsbRepOverList.append(eplst)
                # print(eplst)

        if epsbRepOverList:
            debugPrint('Attention! Overlap Found!', color='red')
            for record in epsbRepOverList:
                print(record)
        else:
            debugPrint('No Overlap Record.', color='green')

    def isStrDate(self, dateStr):
        self.go2MainMenu()
        self.sendString(22, 8, 'red,0' + self.ssrSN + ',' + dateStr)
        self.sendEnter()
        dateGet = self.getString(4, 13, 2) + self.getString(4, 10, 2) + self.getString(4, 7, 2)
        if dateGet == dateStr:
            return True
        else:
            return False

    def typeOfRedArgs(self, argsList):
        """
        dateFrom, dateTo, ssrSN, <s, w, t, b, e>
        """
        # funcArgs for store letter in ['s', 'w', 't', 'b', 'e']
        funcArgs = []
        dateArgs = []
        ssrArgs = []
        funcArgsList = ['S', 'W', 'T', 'B', 'E', 'D']
        for args in argsList:
            args = args.upper()
            if args in funcArgsList:
                funcArgs.append(args)
                continue
            if self.isStrDate(args):
                dateArgs.append(args)
                continue
            if self.transSN2notesID(args):
                ssrArgs.append(args)
                continue
            if args.startswith('C') or args.startswith('0'):
                if self.transSN2notesID(args[1:len(args)]):
                    ssrArgs.append(args[1:len(args)])
                    continue
        dateArgs.sort()
        return {'funcArgs': funcArgs, 'dateArgs': dateArgs, 'ssrArgs': ssrArgs}

    def redPlus(self, redArgs, cutFrom='', cutTo='', teamSSRList='', epsbSSRList=''):
        """
        ['s', 't', 'b', 'e']
        red,s,w,t,b,e,ssr sn,date from,date to',
        <s: summary; w: weekly; t: team UT summary;
        <b: backup to file; e: EPSB overlap verify;
        date format: 200930, ignore 2 date order.
        """
        # fromDate = ''
        # toDate = ''
        saveOrNot = False
        sumOrNot = False
        detailOrNot = False
        todayStr = str(date.today()).replace('20', '', 1).replace('-', '')
        redArgsDict = self.typeOfRedArgs(redArgs)
        print(redArgsDict)
        funcArgs = redArgsDict['funcArgs']
        dateArgs = redArgsDict['dateArgs']
        ssrArgs = redArgsDict['ssrArgs']

        if len(dateArgs) >= 2:
            fromDate = dateArgs[0]
            toDate = dateArgs[1]
        elif len(dateArgs) == 1:
            fromDate = toDate = dateArgs[0]
        elif 'S' in funcArgs or 'T' in funcArgs:
            fromDate = cutFrom
            toDate = cutTo
        else:
            fromDate = toDate = todayStr

        # Debug, delete later.
        # print('From Date: ' + fromDate + '\n' + 'To Date: ' + toDate)

        if len(ssrArgs) >= 2:
            teamSSRList = epsbSSRList = ssrArgs
        elif len(ssrArgs) == 0:
            ssrArgs.append(self.ssrSN)

        # Backup
        if 'B' in funcArgs:
            saveOrNot = True
        # Save
        if 'S' in funcArgs:
            sumOrNot = True
        # Should be single day, if 'D' is specified
        # Detail
        if 'D' in funcArgs:
            detailOrNot = True
        # EPSB Verify
        if 'E' in funcArgs:
            self.teamEPSBCheck(epsbSSRList, fromDate, toDate)
            return True
        # Team
        if 'T' in funcArgs:
            self.getUTSum4Team(teamSSRList, fromDate, toDate)
            return True

        for ssr in ssrArgs:
            self.readTimeReport(ssr, fromDate, toDate, save=saveOrNot, utSum=sumOrNot, prtDetail=detailOrNot)

    def whoIs(self, wArgs):
        pass


class retainp(x3270):
    def __init__(self, host, username, password):
        super(retainp, self).__init__(host, backupHost='')


if __name__ == '__main__':
    pass


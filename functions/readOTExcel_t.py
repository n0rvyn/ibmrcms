from functions import read_config
import os
from openpyxl import load_workbook
from datetime import date
import datetime
from functions import transform_rcms_date
import time

ot_files_dir = ''
ot_files = []
ot_items = []

try:
    ot_files_dir = read_config.ot_file_loc
    if ot_files_dir == '':
        raise RuntimeWarning('OT Files Not Defined!')
    for file in os.listdir(ot_files_dir):
        if file.endswith('xlsx') or file.endswith('xls'):
            if not file.startswith('~') and not file.startswith('.'):
                ot_files.append(file)
except RuntimeError:
    print('Read OT Request Form Location Failed!')
    
try:
    os.chdir(ot_files_dir)
except IsADirectoryError:
    print("'{0}' May not Exist.".format(ot_files_dir))

for file in ot_files:
    workbook = load_workbook(file, read_only=True, data_only=True)
    sheet = workbook.get_sheet_by_name('Sheet1')
    max_row = sheet.max_row
    row = 3
    while True:
        row += 1
        if row > max_row:
            break
        if not str(sheet.cell(row, 6).value).startswith('20'):
            continue
        ot_date = sheet.cell(row, 6).value
        ot_start_time = sheet.cell(row, 7).value
        ot_end_time = sheet.cell(row, 8).value
        ot_time = sheet.cell(row, 9).value
        ot_customer = sheet.cell(row, 4).value
        ot_content = sheet.cell(row, 5).value
        ot_items.append([ot_date, ot_start_time, ot_end_time, ot_time, ot_customer, ot_content])

# Change time format
# 18:00 to 180, 18:30 to 185
for line in ot_items:
    try:
        line[0] = line[0].strftime('%y%m%d')
        line[1] = line[1].strftime('%H') + str(int(line[2].strftime('%M')) // 6)
        line[2] = line[2].strftime('%H') + str(int(line[2].strftime('%M')) // 6)
        if line[2] == '000':
            line[2] = '240'
        if type(line[3]) == datetime.time:
            line[3] = '{:0>3d}'.format(int(line[3].strftime('%H')) * 10)
        else:
            raise ValueError('Wrong Format of Date in OT Excel!')
    except ValueError:
        print('Wrong Format of Date in OT Excel!')
        
ot_items.sort()


def get_ot_cycle(ot_month=''):
    try:
        int(ot_month)
    except ValueError:
        print('Wrong Month Format!')
    
    cur_year = int(date.today().strftime('%y'))
    cur_month = int(date.today().strftime('%m'))
    
    if not ot_month:
        ot_month = cur_month
    
    try:
        int(ot_month)
    except ValueError:
        print("'{0}' is Not Number!".format(ot_month))
    
    if int(ot_month) > 12 or int(ot_month) < 1:
        print('Wrong Month!')
        return None
    '''
    If current month is 2 ad ot_month input '6'
    Then program think this month is in the last year
    '''
    ot_month = int(ot_month)
    
    if 2 <= ot_month <= cur_month:
        last_month = ot_month - 1
        last_year = cur_year
    elif ot_month == 1:
        last_year = cur_year - 1
        last_month = ot_month + 11
    else:
        last_year = cur_year - 1
        last_month = ot_month - 1
        cur_year -= 1
    
    sta_date = '{:0>2d}{:0>2d}{:0>2d}'.format(last_year, last_month, 21)
    end_date = '{:0>2d}{:0>2d}{:0>2d}'.format(cur_year, ot_month, 20)
    ot_cycle = [str(sta_date), str(end_date)]
    
    return ot_cycle
    

def get_ot_in_cycle(start_stop_date=[]):
    sta_date = start_stop_date[0]
    stp_date = start_stop_date[1]
    ot_items_in_cycle = []
    
    for item in ot_items:
        if sta_date <= item[0] <= stp_date:
            if item not in ot_items_in_cycle:
                ot_items_in_cycle.append(item)
            
    return ot_items_in_cycle


def get_ot_date_list(start_stop_date=[]):
    if not start_stop_date:
        return None
    
    ot_lists = get_ot_in_cycle(start_stop_date)
    ot_set = set([])
    
    for ot_list in ot_lists:
        ot_set.add(ot_list[0])
    
    ot_date_list = list(ot_set)
    ot_date_list.sort()
    
    return ot_date_list


def get_one_day_ot_time(overtime_date):
    try:
        int(overtime_date)
    except ValueError:
        print('Wrong Time Report Date Format!')
    
    ot_month = overtime_date[2:4]
    ot_day = overtime_date[4:6]
    
    cur_month = int(date.today().strftime('%m'))
    ot_month = int(ot_month)
    if int(ot_day) > 20:
        ot_month_tmp = ot_month + 1
    else:
        ot_month_tmp = ot_month
    
    ot_cycle = get_ot_cycle(ot_month_tmp)
    ot_lists = get_ot_in_cycle(ot_cycle)
    ot_date_list = get_ot_date_list(ot_cycle)
    
    tmp_date = '20' + str('{:0>2d}'.format(ot_month)) + str(ot_day)
    ot_time_set = set([])
    for ot_list in ot_lists:
        if ot_list[0] == tmp_date:
            ot_time_set = ot_time_set | transform_rcms_date.sta_trans_to_set(ot_list[1], ot_list[2])
            # ot_time_set.add(transform_rcms_date.sta_trans_to_set(ot_list[1], ot_list[2]))
            '''
            time_point = 0
            while int(ot_list[1]) + time_point <= int(ot_list[2]):
                ot_time_set.add('{:0>3d}'.format(int(ot_list[1]) + time_point))
                time_point += 5
            '''
    
    ot_time_list = list(ot_time_set)
    ot_time_list.sort()
    
    print(ot_time_list)
    return ot_time_set


